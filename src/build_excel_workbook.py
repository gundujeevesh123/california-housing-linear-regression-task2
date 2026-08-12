"""Build Task2_Predictions.xlsx as three self-contained chart sheets plus
README and Config. Each chart sheet holds its own data table and the chart
built from that table, so opening the file makes it immediately obvious
which numbers produced which visual.

Sheets
------
  README                  How to read and verify the workbook.
  1_Actual_vs_Predicted   Dataset 1 (Actual + Decision-Tree Predicted, 4,128 rows)
                          → ScatterChart with red y = x reference line.
  2_Residuals             Dataset 2 (Predicted + Residual, 4,128 rows)
                          → ScatterChart with red y = 0 reference line.
  3_Model_Comparison      Dataset 3 (3-row Model / RMSE / R² / Rank table
                          computed with LIVE Excel formulas)
                          → BarChart of RMSE and R² per model.
  Config                  Exact run configuration and library versions.

The predictions are recomputed with scikit-learn (random_state=42, same
80/20 split, StandardScaler fit on training only) so the workbook stands
on its own without any other project files.
"""

from __future__ import annotations

import platform
import sys
from pathlib import Path

import joblib
import matplotlib
import numpy as np
import openpyxl
import pandas as pd
import sklearn
from openpyxl.chart import BarChart, Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sklearn.datasets import fetch_california_housing
from sklearn.linear_model import LinearRegression, Ridge
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.tree import DecisionTreeRegressor

# --------------------------------------------------------------------------
TASK2 = Path(__file__).resolve().parents[1]
XLSX = TASK2 / "outputs" / "Task2_Predictions.xlsx"
XLSX.parent.mkdir(parents=True, exist_ok=True)

FONT_NAME = "Arial"
RANDOM_STATE = 42

# --------------------------------------------------------------------------
# Recompute the pipeline (identical to the notebook)
# --------------------------------------------------------------------------
bundle = fetch_california_housing(as_frame=True)
df = pd.concat([bundle.data, bundle.target.rename("HousePrice")], axis=1)
X = df.drop(columns=["HousePrice"])
y = df["HousePrice"]

X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.2, random_state=RANDOM_STATE,
)
scaler = StandardScaler().fit(X_train)
X_train_s = scaler.transform(X_train)
X_test_s = scaler.transform(X_test)

pred_lr = LinearRegression().fit(X_train_s, y_train).predict(X_test_s)
pred_rd = Ridge(alpha=1.0).fit(X_train_s, y_train).predict(X_test_s)
pred_dt = DecisionTreeRegressor(max_depth=5, random_state=RANDOM_STATE).fit(
    X_train_s, y_train).predict(X_test_s)

y_test_arr = y_test.to_numpy()
N = len(y_test_arr)

# --------------------------------------------------------------------------
# Shared style helpers
# --------------------------------------------------------------------------
thin = Side(border_style="thin", color="AAAAAA")
box  = Border(top=thin, bottom=thin, left=thin, right=thin)
header_fill = PatternFill("solid", fgColor="E6ECF5")
winner_fill = PatternFill("solid", fgColor="DAF0DA")


def header(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(name=FONT_NAME, size=10, bold=True)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = box


def note(ws, coord: str, text: str) -> None:
    ws[coord] = text
    ws[coord].font = Font(name=FONT_NAME, size=9, italic=True)


def red_line(width: float = 1.5) -> GraphicalProperties:
    return GraphicalProperties(
        ln=LineProperties(w=int(width * 12700), solidFill="FF0000"))


def no_line() -> GraphicalProperties:
    return GraphicalProperties(ln=LineProperties(noFill=True))


wb = openpyxl.Workbook()


def setup_landscape_fit(ws) -> None:
    """Make a large sheet print in landscape, fit-to-page — used by chart sheets."""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3

# --------------------------------------------------------------------------
# README
# --------------------------------------------------------------------------
ws = wb.active
ws.title = "README"

readme = [
    ("AI/ML Task 2 — Model Comparison Workbook", True, 14),
    ("", False, 11),
    ("How this workbook is organised", True, 12),
    ("Three chart sheets follow — each one is self-contained:", False, 10),
    ("  • Sheet 1 (1_Actual_vs_Predicted)   Dataset in columns A–B, ",
     False, 10),
    ("      chart on the right. 4,128 test rows.", False, 10),
    ("  • Sheet 2 (2_Residuals)             Dataset in columns A–B, ",
     False, 10),
    ("      chart on the right. Residual = Actual − Predicted (formula in col B).",
     False, 10),
    ("  • Sheet 3 (3_Model_Comparison)      Dataset in cells A1:D4, ",
     False, 10),
    ("      chart on the right. RMSE / R² are Excel formulas — click a cell ",
     False, 10),
    ("      to see the formula in the formula bar.", False, 10),
    ("", False, 11),
    ("Nothing to click, drag, or plot manually. Excel/LibreOffice renders the ",
     False, 10),
    ("charts automatically the moment the file is opened. Change a cell in ",
     False, 10),
    ("column A or B on any chart sheet and the corresponding dot on that ",
     False, 10),
    ("sheet's chart moves — impossible for an embedded image.", False, 10),
    ("", False, 11),
    ("Verified test-set results (from executed notebook)", True, 12),
    ("  Model               Test RMSE   Test R²   Rank", True, 10),
    ("  Linear Regression      0.7456    0.5758     3", False, 10),
    ("  Ridge Regression       0.7456    0.5758     2", False, 10),
    ("  Decision Tree          0.7242    0.5997     1  ← best", True, 10),
    ("", False, 11),
    ("Configuration and library versions live on the 'Config' sheet.",
     False, 10),
]
for i, (text, bold, size) in enumerate(readme, 1):
    c = ws.cell(row=i, column=1, value=text)
    c.font = Font(name=FONT_NAME, size=size, bold=bold,
                  family=3 if "Model" in text or "best" in text else 2)
    c.alignment = Alignment(wrap_text=False, vertical="top")
ws.column_dimensions["A"].width = 92

# ==========================================================================
# Sheet 1 — Dataset 1 + Chart 1 (Actual vs Predicted, Decision Tree)
# ==========================================================================
s1 = wb.create_sheet("1_Actual_vs_Predicted")
setup_landscape_fit(s1)
# Only print the region that has the chart — the 4,128 raw rows are for
# viewing/formulas, not for printing.
s1.print_area = "A1:R32"
header(s1["A1"], "Actual")
header(s1["B1"], "Predicted_DecisionTree")
for i in range(N):
    s1.cell(row=i + 2, column=1, value=float(y_test_arr[i])).number_format = "0.0000"
    s1.cell(row=i + 2, column=2, value=float(pred_dt[i])).number_format = "0.0000"
s1.column_dimensions["A"].width = 12
s1.column_dimensions["B"].width = 24
s1.freeze_panes = "A2"

# Reference-line helper table on the same sheet — no separate sheet needed
lo = float(min(y_test_arr.min(), pred_dt.min()))
hi = float(max(y_test_arr.max(), pred_dt.max()))
header(s1["D1"], "yx_X")
header(s1["E1"], "yx_Y")
s1["D2"], s1["E2"] = round(lo, 4), round(lo, 4)
s1["D3"], s1["E3"] = round(hi, 4), round(hi, 4)
for col in "DE":
    s1.column_dimensions[col].width = 10
note(s1, "D5",
     "Two points defining the red diagonal y = x reference line (D2:E2 → D3:E3).")

# Build the chart
c1 = ScatterChart()
c1.title = "Chart 1 — Actual vs Predicted (Decision Tree, test set)"
c1.style = 2
c1.x_axis.title = "Actual House Prices"
c1.y_axis.title = "Predicted House Prices"
c1.x_axis.scaling.min = 0
c1.x_axis.scaling.max = 5.5
c1.y_axis.scaling.min = 0
c1.y_axis.scaling.max = 5.5
c1.legend.position = "t"

x_actual = Reference(s1, min_col=1, min_row=2, max_row=N + 1)
y_dtpred = Reference(s1, min_col=2, min_row=2, max_row=N + 1)
ser = Series(y_dtpred, xvalues=x_actual, title="Test-set predictions")
ser.marker = Marker(symbol="circle", size=4)
ser.marker.spPr = GraphicalProperties(
    ln=LineProperties(noFill=True), solidFill="4C78A8")
ser.graphicalProperties = no_line()
c1.series.append(ser)

xr = Reference(s1, min_col=4, min_row=2, max_row=3)
yr = Reference(s1, min_col=5, min_row=2, max_row=3)
line = Series(yr, xvalues=xr, title="Perfect prediction (y = x)")
line.marker = Marker(symbol="none")
line.graphicalProperties = red_line(1.8)
c1.series.append(line)

c1.width, c1.height = 16, 12
s1.add_chart(c1, "F1")
note(s1, "G22",
     "This chart plots columns A (Actual) vs B (Predicted) with the red y=x reference "
     "from D2:E3. Try editing any value in column B and watch the corresponding dot move.")

# ==========================================================================
# Sheet 2 — Dataset 2 + Chart 2 (Residuals vs Predicted)
# ==========================================================================
s2 = wb.create_sheet("2_Residuals")
setup_landscape_fit(s2)
s2.print_area = "A1:R32"
header(s2["A1"], "Predicted_DecisionTree")
header(s2["B1"], "Residual (=Actual − Predicted)")
# Column A: the same DT predictions
# Column B: FORMULA — residual = actual − predicted; actual pulled from sheet 1
for i in range(N):
    r = i + 2
    s2.cell(row=r, column=1, value=float(pred_dt[i])).number_format = "0.0000"
    s2.cell(row=r, column=2,
            value=f"='1_Actual_vs_Predicted'!A{r}-A{r}").number_format = "0.0000"
s2.column_dimensions["A"].width = 24
s2.column_dimensions["B"].width = 26
s2.freeze_panes = "A2"

# Zero-reference helper
header(s2["D1"], "zero_X")
header(s2["E1"], "zero_Y")
s2["D2"], s2["E2"] = 0.0, 0.0
s2["D3"], s2["E3"] = 5.5, 0.0
for col in "DE":
    s2.column_dimensions[col].width = 10
note(s2, "D5",
     "Two points defining the red y = 0 reference line (D2:E2 → D3:E3).")

c2 = ScatterChart()
c2.title = "Chart 2 — Residuals vs Predicted (Decision Tree, test set)"
c2.style = 2
c2.x_axis.title = "Predicted House Prices"
c2.y_axis.title = "Residual (Actual − Predicted)"
c2.x_axis.scaling.min = 0
c2.x_axis.scaling.max = 5.5
c2.legend.position = "t"

x_p = Reference(s2, min_col=1, min_row=2, max_row=N + 1)
y_r = Reference(s2, min_col=2, min_row=2, max_row=N + 1)
sres = Series(y_r, xvalues=x_p, title="Residual = actual − predicted")
sres.marker = Marker(symbol="circle", size=4)
sres.marker.spPr = GraphicalProperties(
    ln=LineProperties(noFill=True), solidFill="6BAED6")
sres.graphicalProperties = no_line()
c2.series.append(sres)

xz = Reference(s2, min_col=4, min_row=2, max_row=3)
yz = Reference(s2, min_col=5, min_row=2, max_row=3)
zeroline = Series(yz, xvalues=xz, title="Zero reference (y = 0)")
zeroline.marker = Marker(symbol="none")
zeroline.graphicalProperties = red_line(1.5)
c2.series.append(zeroline)

c2.width, c2.height = 16, 12
s2.add_chart(c2, "F1")
note(s2, "G22",
     "Column B is a formula (=Sheet1!A_i − A_i). Change any Predicted value in "
     "column A and both the corresponding residual (col B) and this chart update.")

# ==========================================================================
# Sheet 3 — Dataset 3 + Chart 3 (Model comparison bars)
# ==========================================================================
s3 = wb.create_sheet("3_Model_Comparison")
setup_landscape_fit(s3)
s3.print_area = "A1:R32"

# Data table header
for col, h in enumerate(["Model", "RMSE (test)", "R² (test)", "Rank"], 1):
    header(s3.cell(row=1, column=col), h)

# Cross-sheet references for the RMSE / R² formulas.
# We need Ridge & Linear predictions too; keep them on a hidden helper sheet
# so this sheet stays clean.
helper = wb.create_sheet("_Predictions_Full")
helper.sheet_state = "hidden"
header(helper["A1"], "Actual")
header(helper["B1"], "Linear_Predicted")
header(helper["C1"], "Ridge_Predicted")
header(helper["D1"], "DecisionTree_Predicted")
for i in range(N):
    r = i + 2
    helper.cell(row=r, column=1, value=float(y_test_arr[i])).number_format = "0.0000"
    helper.cell(row=r, column=2, value=float(pred_lr[i])).number_format = "0.0000"
    helper.cell(row=r, column=3, value=float(pred_rd[i])).number_format = "0.0000"
    helper.cell(row=r, column=4, value=float(pred_dt[i])).number_format = "0.0000"

def rmse_formula(pred_col: str) -> str:
    a = f"_Predictions_Full!$A$2:$A${N + 1}"
    p = f"_Predictions_Full!${pred_col}$2:${pred_col}${N + 1}"
    return f"=SQRT(SUMPRODUCT(({a}-{p})^2)/{N})"

def r2_formula(pred_col: str) -> str:
    a = f"_Predictions_Full!$A$2:$A${N + 1}"
    p = f"_Predictions_Full!${pred_col}$2:${pred_col}${N + 1}"
    return f"=1-SUMPRODUCT(({a}-{p})^2)/SUMPRODUCT(({a}-AVERAGE({a}))^2)"

rows = [
    ("Linear Regression", "B"),
    ("Ridge Regression", "C"),
    ("Decision Tree",     "D"),
]
for i, (name, col_letter) in enumerate(rows, 2):
    s3.cell(row=i, column=1, value=name).border = box
    s3.cell(row=i, column=2, value=rmse_formula(col_letter))
    s3.cell(row=i, column=3, value=r2_formula(col_letter))
    s3.cell(row=i, column=4, value=f"=RANK(B{i},$B$2:$B$4,1)")
    s3.cell(row=i, column=2).number_format = "0.0000"
    s3.cell(row=i, column=3).number_format = "0.0000"
    for c in range(1, 5):
        s3.cell(row=i, column=c).border = box
        s3.cell(row=i, column=c).font = Font(name=FONT_NAME, size=10)
    s3.cell(row=i, column=4).alignment = Alignment(horizontal="center")

# Highlight the winning row (Decision Tree, row 4)
for c in range(1, 5):
    s3.cell(row=4, column=c).fill = winner_fill
    s3.cell(row=4, column=c).font = Font(name=FONT_NAME, size=10, bold=True)

for i, w in enumerate([26, 14, 14, 10], 1):
    s3.column_dimensions[get_column_letter(i)].width = w

note(s3, "A7",
     "RMSE and R² are Excel formulas (click cell B2, C2, etc. to see them). "
     "Rank uses =RANK on the RMSE column so it updates automatically if the "
     "underlying predictions change.")
s3.merge_cells("A7:F7")

# Bar chart
c3 = BarChart()
c3.type = "col"
c3.style = 2
c3.title = "Chart 3 — Model comparison (RMSE and R² on test set)"
c3.x_axis.title = "Model"
c3.y_axis.title = "Metric value"
c3.legend.position = "t"

data = Reference(s3, min_col=2, max_col=3, min_row=1, max_row=4)
cats = Reference(s3, min_col=1, max_col=1, min_row=2, max_row=4)
c3.add_data(data, titles_from_data=True)
c3.set_categories(cats)
c3.width, c3.height = 18, 12
s3.add_chart(c3, "F1")

# ==========================================================================
# Config sheet
# ==========================================================================
cfg = wb.create_sheet("Config")
cfg_rows = [
    ("Dataset",             "California Housing via sklearn.datasets.fetch_california_housing(as_frame=True)"),
    ("Target column",       "HousePrice"),
    ("Rows total",          len(df)),
    ("Rows train",          len(X_train)),
    ("Rows test",           len(X_test)),
    ("test_size",           0.2),
    ("random_state",        RANDOM_STATE),
    ("Preprocessing",       "StandardScaler fit on X_train only, then transform train & test"),
    ("Linear Regression",   "LinearRegression()"),
    ("Ridge Regression",    "Ridge(alpha=1.0)"),
    ("Decision Tree",       "DecisionTreeRegressor(max_depth=5, random_state=42)"),
    ("Selection rule",      "argmin(test RMSE); tiebreak argmax(test R²)"),
    ("", ""),
    ("Python",              sys.version.split()[0]),
    ("OS",                  platform.platform()),
    ("scikit-learn",        sklearn.__version__),
    ("pandas",              pd.__version__),
    ("numpy",               np.__version__),
    ("matplotlib",          matplotlib.__version__),
    ("joblib",              joblib.__version__),
    ("openpyxl",            openpyxl.__version__),
]
for i, (k, v) in enumerate(cfg_rows, 1):
    kc = cfg.cell(row=i, column=1, value=k)
    vc = cfg.cell(row=i, column=2, value=v)
    kc.font = Font(name=FONT_NAME, size=11, bold=True)
    vc.font = Font(name=FONT_NAME, size=11)
cfg.column_dimensions["A"].width = 25
cfg.column_dimensions["B"].width = 80

# Final tab order
wb._sheets = [wb["README"],
              wb["1_Actual_vs_Predicted"],
              wb["2_Residuals"],
              wb["3_Model_Comparison"],
              wb["_Predictions_Full"],
              wb["Config"]]

wb.save(XLSX)
print(f"Wrote: {XLSX}  ({XLSX.stat().st_size:,} bytes)")
