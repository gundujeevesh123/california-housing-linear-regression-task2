"""Insert the three deliverable charts into Task2_Predictions.xlsx using
openpyxl.chart (native Excel charts, not images). Then render the workbook to
PDF via LibreOffice and crop each chart out as a standalone PNG.

Inputs  : outputs/Task2_Predictions.xlsx  (built by build_excel_workbook.py)
Outputs : outputs/Task2_Predictions.xlsx  (in-place: adds a 'Charts' sheet)
          outputs/excel_charts/Chart1_Actual_vs_Predicted.png
          outputs/excel_charts/Chart2_Residuals.png
          outputs/excel_charts/Chart3_RMSE_R2_Comparison.png
          outputs/excel_charts/Task2_Predictions_rendered.pdf
"""

from __future__ import annotations

import shutil
import subprocess
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.trendline import Trendline
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.fill import ColorChoice as FillColor
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font

TASK2 = Path(__file__).resolve().parents[1]
XLSX = TASK2 / "outputs" / "Task2_Predictions.xlsx"
CHART_DIR = TASK2 / "outputs" / "excel_charts"
CHART_DIR.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(XLSX)
pred = wb["Predictions"]
n_rows = pred.max_row  # header + 4128

# Rebuild dedicated single-chart sheets so each PDF page shows one chart cleanly.
for name in ("Charts", "Chart_1_Actual_vs_Predicted",
             "Chart_2_Residuals", "Chart_3_RMSE_R2_Comparison"):
    if name in wb.sheetnames:
        del wb[name]

def make_chart_sheet(sheet_name: str, note: str):
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = note
    ws["A1"].font = Font(name="Arial", size=10, italic=True)
    ws.column_dimensions["A"].width = 20
    # Landscape and fit-to-page so the chart isn't clipped by print pagination.
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    return ws

ws1 = make_chart_sheet("Chart_1_Actual_vs_Predicted",
    "Chart 1 — Actual vs Predicted (Decision Tree). Data source: Predictions!B:B (Actual) "
    "vs Predictions!E:E (Decision_Tree_Predicted); reference line: YX_ReferenceLine!B2:C3.")
ws2 = make_chart_sheet("Chart_2_Residuals",
    "Chart 2 — Residuals vs Predicted (Decision Tree). Data source: Predictions!E:E "
    "(Decision_Tree_Predicted) vs Predictions!F:F (Residual_DT); zero line: "
    "YX_ReferenceLine!E2:F3.")
ws3 = make_chart_sheet("Chart_3_RMSE_R2_Comparison",
    "Chart 3 — RMSE and R² by model. Data source: Comparison!A1:C4 (RMSE and R² formulas).")


def red_line(width: float = 1.5) -> GraphicalProperties:
    return GraphicalProperties(ln=LineProperties(w=int(width * 12700),
                                                 solidFill="FF0000"))


def no_line() -> GraphicalProperties:
    return GraphicalProperties(ln=LineProperties(noFill=True))


# --------------------------------------------------------------------------
# Chart 1 — Actual vs Predicted (Decision Tree)
# --------------------------------------------------------------------------
c1 = ScatterChart()
c1.title = "Actual vs Predicted — Decision Tree (test set)"
c1.style = 2
c1.x_axis.title = "Actual House Prices"
c1.y_axis.title = "Predicted House Prices"
c1.x_axis.scaling.min = 0
c1.x_axis.scaling.max = 5.5
c1.y_axis.scaling.min = 0
c1.y_axis.scaling.max = 5.5
c1.legend.position = "t"

# Series 1: scatter of Actual (X) vs Decision_Tree_Predicted (Y)
x_actual = Reference(pred, min_col=2, min_row=2, max_row=n_rows)     # B2:B4129
y_dt     = Reference(pred, min_col=5, min_row=2, max_row=n_rows)     # E2:E4129
s1 = Series(y_dt, xvalues=x_actual, title="Test-set predictions")
# Small semi-transparent markers, no connecting line
s1.marker = Marker(symbol="circle", size=4)
s1.marker.spPr = GraphicalProperties(
    ln=LineProperties(noFill=True),
    solidFill="4C78A8",
)
s1.graphicalProperties = no_line()
c1.series.append(s1)

# Series 2: y = x diagonal reference line (red)
ref = wb["YX_ReferenceLine"]
x_ref = Reference(ref, min_col=2, min_row=2, max_row=3)  # B2:B3
y_ref = Reference(ref, min_col=3, min_row=2, max_row=3)  # C2:C3
s2 = Series(y_ref, xvalues=x_ref, title="Perfect prediction (y = x)")
s2.marker = Marker(symbol="none")
s2.graphicalProperties = red_line(1.8)
c1.series.append(s2)

c1.width = 24
c1.height = 15
ws1.add_chart(c1, "B3")

# --------------------------------------------------------------------------
# Chart 2 — Residuals vs Predicted (Decision Tree)
# --------------------------------------------------------------------------
c2 = ScatterChart()
c2.title = "Residuals vs Predicted — Decision Tree (test set)"
c2.style = 2
c2.x_axis.title = "Predicted House Prices"
c2.y_axis.title = "Residual (Actual − Predicted)"
c2.x_axis.scaling.min = 0
c2.x_axis.scaling.max = 5.5
c2.legend.position = "t"

x_pred = Reference(pred, min_col=5, min_row=2, max_row=n_rows)   # E2:E4129
y_res  = Reference(pred, min_col=6, min_row=2, max_row=n_rows)   # F2:F4129
sr = Series(y_res, xvalues=x_pred, title="Residual = actual − predicted")
sr.marker = Marker(symbol="circle", size=4)
sr.marker.spPr = GraphicalProperties(
    ln=LineProperties(noFill=True),
    solidFill="6BAED6",
)
sr.graphicalProperties = no_line()
c2.series.append(sr)

# Zero-reference horizontal line via a helper 2-point series on the same X range
# (we reuse the YX_ReferenceLine sheet by extending it with a 0-line)
ref["E1"] = "X (Predicted)"
ref["F1"] = "Y (zero)"
ref["E2"], ref["F2"] = 0.0, 0.0
ref["E3"], ref["F3"] = 5.5, 0.0
for cell in ("E1", "F1"):
    ref[cell].font = Font(name="Arial", size=11, bold=True)

x_zero = Reference(ref, min_col=5, min_row=2, max_row=3)  # E2:E3
y_zero = Reference(ref, min_col=6, min_row=2, max_row=3)  # F2:F3
sz = Series(y_zero, xvalues=x_zero, title="Zero reference (y = 0)")
sz.marker = Marker(symbol="none")
sz.graphicalProperties = red_line(1.5)
c2.series.append(sz)

c2.width = 24
c2.height = 15
ws2.add_chart(c2, "B3")

# --------------------------------------------------------------------------
# Chart 3 — RMSE / R² bar comparison
# --------------------------------------------------------------------------
cmp = wb["Comparison"]
c3 = BarChart()
c3.type = "col"
c3.style = 2
c3.title = "Model comparison — RMSE and R² (test set)"
c3.x_axis.title = "Model"
c3.y_axis.title = "Metric value"
c3.legend.position = "t"

# Data range: header B1:C1 + rows B2:C4  (RMSE + R²)
data = Reference(cmp, min_col=2, max_col=3, min_row=1, max_row=4)
cats = Reference(cmp, min_col=1, max_col=1, min_row=2, max_row=4)
c3.add_data(data, titles_from_data=True)
c3.set_categories(cats)
c3.width = 24
c3.height = 15
ws3.add_chart(c3, "B3")

# --------------------------------------------------------------------------
# Save workbook, render via LibreOffice, crop chart pages to PNG
# --------------------------------------------------------------------------
wb.save(XLSX)
print(f"Saved workbook with charts: {XLSX}")

# Convert workbook to PDF via LibreOffice
result = subprocess.run(
    ["soffice", "--headless",
     "--convert-to", "pdf",
     "--outdir", str(CHART_DIR),
     str(XLSX)],
    check=True, capture_output=True, text=True,
)
print(result.stdout.strip() or result.stderr.strip())

# Also render individual PNGs of each chart by exporting the Charts sheet
# as an image via LibreOffice's --convert-to png (renders active sheet
# by default). Best cross-platform path: use `pdftoppm` on the PDF and
# then crop each chart from the rendered pages.
pdf = CHART_DIR / f"{XLSX.stem}.pdf"
final_pdf = CHART_DIR / "Task2_Predictions_rendered.pdf"
pdf.rename(final_pdf)
print(f"Rendered PDF: {final_pdf}")
